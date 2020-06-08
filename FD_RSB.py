import cv2
import os
import numpy as np
import sqlite3
from PIL import Image,ImageTk
from tkinter import *
import turtle
import time
from tkinter import ttk
from tkcalendar import Calendar, DateEntry
import re
import openpyxl
import pandas as pd
import winsound

def beep():
    winsound.PlaySound('bp.wav', winsound.SND_FILENAME)

def trainer():
    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        detector = cv2.CascadeClassifier(r"haarcascade\haarcascade_frontalface_default.xml")

        def getImagesAndLabels(path):
            imagePaths = [os.path.join(path, f) for f in os.listdir(path)]
            # print(imagePaths)

            faceSamples = []

            ids = []

            for imagePath in imagePaths:
                PIL_img = Image.open(imagePath).convert('L')
                # PIL_img = cv2.imread(imagePath,cv2.IMREAD_GRAYSCALE)

                img_numpy = np.array(PIL_img, 'uint8')
                # facenp = np.array(PIL_img, 'uint8')
                # print(img_numpy)

                id = int(os.path.split(imagePath)[-1].split(".")[1])
                # faceSamples.append(facenp)
                # ids.append(id)
                # print(faceSamples)
                # print(id)
                faces = detector.detectMultiScale(img_numpy)
                # print(faces)

                for (x, y, w, h) in faces:
                    faceSamples.append(img_numpy[y: y + h, x: x + w])

                    ids.append(id)
                # print(faceSamples)
                # print(ids)
            return faceSamples, ids

        faces, ids = getImagesAndLabels('data1')

        recognizer.train(faces, np.array(ids))

        # Save the model into trainer.yml
        # assure_path_exists('trainer1/')
        recognizer.save('trainer1/trainer1.yml')
        Label(admin_screen, text="Training Completed", fg="black", bg='sky blue', font=("Georgia", 12, 'bold')).grid(
            row=10, column=2)
        # gui.update()
    except:
        print('INFO...[ Images not available]')


def imgdetector():




    try:
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.read('trainer1/trainer1.yml')
        cascadePath = (r"haarcascade\haarcascade_frontalface_default.xml")
        faceCascade = cv2.CascadeClassifier(cascadePath)

        def getp(Id):
            conn = sqlite3.connect("datacollector.db")
            cmd = "SELECT * FROM COMPANY WHERE ID=" + str(Id)
            cursor = conn.execute(cmd)
            pr = None
            for row in cursor:
                pr = row
            conn.close()
            return pr

        cam = cv2.VideoCapture(0)
        font = cv2.FONT_HERSHEY_COMPLEX
        while True:
            ret, im = cam.read()
            gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)

            faces = faceCascade.detectMultiScale(gray, 1.3, 5)
            for (x, y, w, h) in faces:
                cv2.line(im, (x, y - 30), (x + (w // 5), y - 30), (0, 0, 0), 2)
                cv2.line(im, (x, y - 30), (x, y + (h // 5) - 30), (0, 0, 0), 2)
                # cv2.line(im, (x+int(w/2)-30 , y),(x+int(w/2)+30 ,y), (0,0,0),2)
                # cv2.line(im, (x+int(w/2),y-5),(x+int(w/2),y+5),(0,0,0),1)

                cv2.line(im, (x, y + h), (x + (w // 5), y + h), (0, 0, 0), 2)
                cv2.line(im, (x, y + h), (x, y + h - (h // 5)), (0, 0, 0), 2)
                # cv2.line(im, (x ,y + ( h // 5 ) +15), (x,y +h-50),(0,0,0),2)
                # cv2.line(im, (x-5, y+int(h/2)),(x+5,y+int(h/2)),(0,0,0),1)

                cv2.line(im, (x + w, y - 30), (x + w - (w // 5), y - 30), (0, 0, 0), 2)
                cv2.line(im, (x + w, y - 30), (x + w, y + (h // 5) - 30), (0, 0, 0), 2)
                # cv2.line(im,(x+w , y + ( h // 5 ) + 15) ,(x+w,y+h-50),(0,0,0),2)
                # cv2.line(im, (x+w-5,y+int(h/2)),(x+w+5,y+int(h/2)),(0,0,0),1)

                cv2.line(im, (x + w, y + h), (x + w, y + h - (h // 5)), (0, 0, 0), 2)
                cv2.line(im, (x + w, y + h), (x + w - (w // 5), y + h), (0, 0, 0), 2)
                # cv2.line(im, (x+int(w/2)-30,y+h),(x+int(w/2)+30,y+h),(0,0,0),2)
                # cv2.line(im, (x+int(w/2),y+h-5),(x+int(w/2),y+h+5),(0,0,0),1)

                Id, conf = recognizer.predict(gray[y:y + h, x:x + w])
                pr = getp(Id)
                if (pr != None):
                    loc = time.strftime("%H%M%S")
                    tn = time.strftime("%H:%M:%S")
                    dt = time.strftime("%Y-%m-%d")

                    if (conf < 85):

                        cv2.rectangle(im, (x - 20, y + h + 10), (x + w + 22, y + h + 60), (0, 0, 0), -1)
                        cv2.putText(im, str(pr[1]), (x + 5, y + h + 50), font, 1, (255, 255, 255), 1)
                        cv2.putText(im, str(pr[2]), (x + 5, y + h + 90), font, 1, (255, 255, 255), 1)
                        cv2.putText(im, "Press q To Stop", (x, y - 100), font, 1, (0, 0, 0), 2)


                        conn = sqlite3.connect("datacollector.db")
                        conn.execute('INSERT INTO LOGIN(ID,NAME,Login_Time,Login_Date) VALUES (?, ?, ?, ?)',(pr[0], str(pr[1]), tn, dt))

                        conn.commit()
                        conn.close()
                        #time.sleep(1)
                        # cv2.putText(im, str(conf), (x, y - 40), font, 1, (0, 0, 0), 3)
                    else:

                        # localtime = time.asctime(time.localtime(time.time()))
                        conn = sqlite3.connect("datacollector.db")
                        conn.execute('INSERT INTO LOGIN(ID,NAME,Login_Time,Login_Date) VALUES (?, ?, ?, ?)',
                                     ('Unrecognized', 'Unrecognized', tn, dt))
                        # time.sleep(2)
                        conn.commit()
                        conn.close()

                        cv2.imwrite(r"Unrecognized/Intruder." + str(loc) + ".jpg", im)
                        cv2.waitKey(100)

                        cv2.putText(im, "Not recognized", (x, y - 40), font, 1, (0, 0, 0), 2)
                        cv2.putText(im, "Press q To Stop", (x, y - 100), font, 1, (0, 0, 0), 2)

                        beep()
                        # localtime = time.asctime(time.localtime(time.time()))

                        # cv2.putText(im, str(conf), (x, y - 70), font, 1, (0, 0, 0), 3)
                else:
                    cv2.putText(im, "[INFO]..Dataset Empty", (x, y - 40), font, 1, (0, 0, 0), 2)

            cv2.imshow('PREDICTING', im)

            if cv2.waitKey(10) & 0xFF == ord('q'):
                cam.release()
                break
            if cv2.waitKey(10) & 0xFF == ord('Q'):
                cam.release()
                break

        cam.release()
        cv2.destroyAllWindows()

    except:
        print('INFO...[Data not available]')


turtle.setup(900,500)
t = turtle.Turtle()
screen = turtle.Screen()
screen.title("Face Recognition")
screen.bgpic("t.png")
t.hideturtle()

t.penup()
t.goto(0,110)
t.pendown()
t.pencolor("black")
t.write("Face Recognition",move='true',align="center",font=("Georgia",50))
t.penup()
t.goto(-20,70)
t.pendown()
t.write("Developed By : Ahtesham Zaidi",align="left",font=("georgia",13,"normal",'italic'))

t.penup()
t.goto(-207,-130)
t.pendown()
t.write("Loading .......",align='right',font=("Ariel",13,'normal','italic'))

t.pensize(3)
t.penup()
t.goto(-300,-100)
t.pendown()
t.pencolor("white")
t.forward(200)
time.sleep(1)
t.forward(150)
time.sleep(1)
t.forward(200)

time.sleep(2)
t.forward(190)
time.sleep(1)
turtle.bye()

def Exit():
    return exit()

def admin():
    global admin_screen
    admin_screen = Toplevel(gui)
    admin_screen.configure(bg='white')
    admin_screen.title("ADMIN WINDOW")
    def imgcreate():
        fd = cv2.CascadeClassifier(r"haarcascade\haarcascade_frontalface_default.xml")
        cam = cv2.VideoCapture(0)

        def insert(id, name, gender):
            conn = sqlite3.connect("datacollector.db")
            cmd = 'SELECT * FROM COMPANY WHERE ID ={}'.format(id)
            cursor = conn.execute(cmd)
            ire = 0
            for row in cursor:
                ire = 1

            if (ire == 1):
                conn.execute('UPDATE COMPANY SET NAME = (?) WHERE ID =(?)', (name, id))
                conn.execute('UPDATE COMPANY SET GENDER = (?) WHERE ID =(?)', (gender, id))
            else:
                conn.execute('INSERT INTO COMPANY (ID,NAME,Gender) VALUES (?, ?, ?)', (id, name, gender))
            conn.execute(cmd)
            conn.commit()
            conn.close()

        id = ent.get()

        name = ent1.get()
        if (var.get() == 1):
            gender = 'M'
        elif (var.get() == 2):
            gender = 'F'
        else:
            gender = 'Gender (?)'
        insert(id, name, gender)
        SampleNum = 0
        while (True):
            font = cv2.FONT_HERSHEY_COMPLEX
            ret, img = cam.read()
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = fd.detectMultiScale(gray, 1.3, 5)
            for (x, y, w, h) in faces:
                cv2.rectangle(img, (x, y-35), (x + w  , y + h+20), (255, 255, 255), 1)
                cv2.putText(img, "Move Your Head", (x, y - 100), font, 1, (0, 0, 0), 2)
                SampleNum = SampleNum + 1
                cv2.imwrite(r"data1/User." + id + "." + str(SampleNum) + ".jpg", gray[y:y + h, x:x + w])
                cv2.waitKey(100)
            cv2.imshow("Creating Dataset", img)
            cv2.waitKey(1)
            if (SampleNum == 40):
                cam.release()
                break
        cam.release()
        cv2.destroyAllWindows()
        Label(admin_screen, text="Data Loaded", fg='black',bg='sky blue', font=("Georgia", 12,'bold')).grid(row=7, column=2)
        #gui.update()




    label1 = Label(admin_screen, text="Enter your Id", font=("Georgia", 12), bg='white', fg='black',activebackground='light green')
    label2 = Label(admin_screen, text="Enter your Name ", font=("Georgia", 12), bg='white', fg='black')
    lab2 = Label(admin_screen, text="Select Gender", font=("Georgia", 12), bg='white', fg='black')
    ent = Entry(admin_screen, font=('Ariel', 10), bg='light blue', bd=4)
    ent1 = Entry(admin_screen, font=('Ariel', 10), bg='light blue', bd=4)

    var = IntVar()
    R1 = Radiobutton(admin_screen, text="MALE", variable=var, value=1, bg='white',bd=4,font=("Georgia",8,'bold'))
    R1.grid(row=8, column=1)

    R2 = Radiobutton(admin_screen, text="FEMALE", variable=var, value=2, bg='white',bd=4,font=("Georgia",8,'bold'))
    R2.grid(row=8, column=2)

    # ent2 = Entry(gui)
    # equation.set('enter your expression')
    Label(admin_screen, width=110, bg='blue').grid(row=5, columnspan=3)
    Label(admin_screen, width=110, bg='blue').grid(row=9, columnspan=3)
    label1.grid(row=6, column=0)
    lab2.grid(row=8, column=0)

    ent.grid(row=6, column=1)
    label2.grid(row=7, column=0)
    ent1.grid(row=7, column=1)

    Label(admin_screen, text="Train Dataset", font=('Georgia', 12), fg='black', bg='white').grid(row=10, column=0)
    button2 = Button(admin_screen, text='TRAIN', command=trainer,bd=4,font=("Calibri", 13))
    button2.grid(row=10, column=1)

    button1 = Button(admin_screen,bd=4,text='SAVE', command=imgcreate,font=("Calibri", 13))
    button1.grid(row=6, column=2)
    Label(admin_screen, text="Face Recognition Setup", font=("Georgia", 22), bg="white", fg="black").grid(row=0, column=1)
    Label(admin_screen, image=image1, height=160, width=180).grid(row=1, column=0)


    text2 = Label(admin_screen, image=lb9, height=160, width=250)

    text2.grid(row=1, column=1)

    #btn = Button(gui, image=rg1, command=reg, bd=4).grid(column=2, row=1)

    Label(admin_screen,image=rg1, height=170, width=190).grid(row=1, column=2)

    Label(admin_screen, width=110, height=1, bg='indigo').grid(row=2, columnspan=3)

    button3 = Button(admin_screen, text='DETECT',  command=imgdetector, bd=4,font=("Calibri", 13))
    button3.grid(row=3, column=1)

    Label(admin_screen, height=3, bg='white').grid(row=3, column=2)

    Label(admin_screen, height=3, bg='white').grid(row=11, column=1)

    Button(admin_screen, text= 'Generate Log Sheet',width=15,font=("Calibri", 13),command=login_table, bd=4).grid(row=11, column=2)

def login_table ():
    try:
        top = Toplevel(gui)
        # top.geometry("100x100")
        top.configure(bg='white')
        top.title("GENERATE LOG SHEET")

        def gen():
            def genrate(d, d1, t, t1):

                # loc = time.strftime("%H%M%S")

                wb = openpyxl.Workbook()

                # Get workbook active sheet
                # from the active attribute
                sheet = wb.active

                conn = sqlite3.connect('datacollector.db')
                curs = conn.cursor()
                curs.execute(
                    "SELECT * FROM LOGIN WHERE ID = 'Unrecognized' AND  Login_Date >= (?) AND Login_Date <= (?) AND Login_Time >= (?) AND Login_Time <=(?)",
                    (d, d1, t, t1))
                mysel = curs.execute(
                    "SELECT * FROM LOGIN WHERE ID = 'Unrecognized' AND  Login_Date >= (?) AND Login_Date <= (?) AND Login_Time >= (?) AND Login_Time <=(?)",
                    (d, d1, t, t1))

                # col_names = [cn[0] for cn in curs.description]
                # str(col_names)
                # rows = curs.fetchmany()

                # x = PrettyTable(col_names)
                # x.align[col_names[1]] = "l"
                # x.align[col_names[2]] = "r"
                # x.padding_width = 1
                # for row in rows:
                # add_row(row)
                # print(row)
                # print(x)
                # tabstring = x.get_string()
                # worksheet.write(tabstring)

                sheet.cell(row=1, column=1).value = "ID"
                sheet.cell(row=1, column=2).value = "Name"
                sheet.cell(row=1, column=3).value = "Login_Time"
                sheet.cell(row=1, column=4).value = "Login_Date"
                sheet.cell(row=1, column=5).value = "File Path"
                sheet.cell(row=1, column=7).value = "IMAGE"

                for i, row in enumerate(mysel):

                    for j, it in enumerate(row):
                        sheet.cell(row=i + 2, column=j + 1).value = it

                wb.save("F:\image recognition\LogSheet\Log_sheet.xlsx")

                # Get workbook active sheet
                # from the active attribute
                sheet = wb.active
                # print(df.Login_Time)

                sheet.cell(row=1, column=1).value = "ID"
                sheet.cell(row=1, column=2).value = "Name"
                sheet.cell(row=1, column=3).value = "Login_Time"
                sheet.cell(row=1, column=4).value = "Login_Date"
                sheet.cell(row=1, column=5).value = "File Path"
                sheet.cell(row=1, column=7).value = "IMAGE"
                df = pd.read_excel("F:\image recognition\LogSheet\Log_sheet.xlsx")
                end = len(df)
                for i in range(0, end):
                    sp = df.iloc[i, 2]
                    cleanString = re.sub('\W+', '', sp)
                    sheet.cell(row=i + 2, column=5).value = 'F:\image recognition\\Unrecognized\Intruder.' + str(
                        cleanString) + '.jpg'
                    sheet.cell(row=i + 2, column=6).value = 'INTRUDER' + str(i + 1)
                    a = 'E' + str(i + 2)
                    b = 'F' + str(i + 2)
                    sheet.cell(row=i + 2, column=7).value = '=HYPERLINK(' + str(a) + ',' + str(b) + ')'

                for i, row in enumerate(mysel):

                    for j, it in enumerate(row):
                        sheet.cell(row=i + 2, column=j + 1).value = it

                wb.save("F:\image recognition\LogSheet\Log_sheet.xlsx")

                # output = open("unrecognized_details.txt", "w")
                # output.write("Login Data")
                # output.write(tabstring)
                # output.close()
                conn.commit()
                conn.close()

            d = cal.selection_get()
            d1 = cal1.selection_get()
            t = tent.get()
            t1 = tent1.get()

            genrate(d, d1, t, t1)





        Label(top,text="Generate Log Sheet", bg="blue", width="100", height="2", font=("Calibri", 13)).grid(columnspan=3,row=0)
        Label(top, text="DATE FROM :", bg="blue", font=("Calibri", 13)).grid(pady=10, row=1, column=0)
        cal = Calendar(top, font="Arial 14", selectmode='day')
        cal.grid(pady=10, row=2, padx=10, column=0)

        # ttk.Button(top, text="ok", command=print_sel).pack()
        Label(top, text="DATE TO :", bg="blue", font=("Calibri", 13)).grid(pady=10, row=1, column=2)
        cal1 = Calendar(top, font="Arial 14", selectmode='day')
        cal1.grid(pady=10, padx=10, row=2, column=2)

        Label(top, text="TIME FROM :", bg="blue", font=("Calibri", 13)).grid(pady=10, column=0, row=3)
        tent = Entry(top, font=('Ariel', 10), bg='light blue', bd=4)
        tent.grid(pady=10, column=0, row=4, padx=10)
        Label(top, text="TIME TO :", bg="blue", font=("Calibri", 13)).grid(pady=10, column=2, row=3)
        tent1 = Entry(top, font=('Ariel', 10), bg='light blue', bd=4)
        tent1.grid(pady=10, column=2, row=4, padx=10)

        # ttk.Button(top, text="ok", command=print_sel).pack()

        Button(top, text='Generate', width=7, height=1,bd=4, command=gen,font=("Calibri", 13,'bold')).grid(column=1, row=5, pady=10)

    except:
        print('INFO...[Something went wrong]')

def mainwin():
    global gui,image1,imab,ex1,de1,sa1,rg1,lb9

    gui = Tk()

    #gui.configure(bg='blue')
    gui.title("Face Recognition Setup")
    gui.geometry("300x250")
    # gui.geometry('300x200')

    # expression_field = Entry(gui, textvariable=equation)
    # expression_field.grid(columnspan=4, ipadx=70)

    image2 = Image.open('us.jpg')
    image1 = ImageTk.PhotoImage(image2)
    bi = Image.open('tra.jpeg')
    imab = ImageTk.PhotoImage(bi)

    ex = Image.open('exit.jpg')
    ex1 = ImageTk.PhotoImage(ex)

    de = Image.open('det.jpeg')
    de1 = ImageTk.PhotoImage(de)

    sa = Image.open('save.jpeg')
    sa1 = ImageTk.PhotoImage(sa)

    rg = Image.open('rg.png')
    rg1 = ImageTk.PhotoImage(rg)

    lb = Image.open('label.jpg')
    lb9 = ImageTk.PhotoImage(lb)
    Label(text="Select Your Choice", bg="blue", width="300", height="2", font=("Calibri", 13)).pack()

    Button(gui,text="ADMIN",width=30,height=2,command=admin).pack(pady=13)
    Button(gui, text="USER", width=30, height=2, command=imgdetector,).pack( pady=13)



    '''mb = Menubutton(gui, text='Select',   width=6, height=1, bd=6)
    mb.winfo_geometry()
    mb.menu = Menu(mb)
    mb['menu'] = mb.menu

    mb.menu.add_command(label='Admin', command=admin)
    mb.menu.add_command(label='User', command=imgdetector)
    
    Label(gui, width=40, height=1, bg='indigo').grid(row=1, columnspan=3)

    mb.grid(row=2, column=1, pady=5)'''



    Button(gui, text="Exit", command=Exit).pack()
    gui.mainloop()

mainwin()

